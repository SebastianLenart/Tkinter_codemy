from tkinter import *
from tkinter import messagebox

from PyInstaller.utils.hooks import collect_all
from openpyxl.workbook import Workbook
from openpyxl import load_workbook


root = Tk()
root.title("resize")
root.geometry("300x300")


def select(e):
    my_label.config(text=my_listbox.get(ANCHOR))

my_list = ["One", "Two", "Three"]
my_listbox = Listbox(root, width=45)
my_listbox.pack()

wb = load_workbook("name_color.xlsx") # tutaj sciezka do .xlsx
ws = wb.active
col_a = ws["A"]
col_b = ws["B"]

for item in col_a:
    my_listbox.insert(END, item.value)

my_label = Label(root, text="Select item..")
my_label.pack()

my_listbox.bind("<ButtonRelease-1", select)

root.mainloop()
