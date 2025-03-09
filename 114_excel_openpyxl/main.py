from tkinter import *
from tkinter import filedialog
from tkinter import font
from tkinter import font
from tkinter import messagebox
from tkinter import colorchooser
import os, sys
from openpyxl import Workbook
from openpyxl import load_workbook

root = Tk()
root.title("resize")
root.geometry("600x480")

wb = Workbook()
wb = load_workbook("pizza.xlsx")
ws = wb.active

column_a = ws["A"]
column_b = ws["B"]

def get_a():
    list = ""
    for cell in column_a:
        print(cell.value)
        list = f"{list + str(cell.value)} \n"
    label_a.config(text=list)

def get_b():
    list = ""
    for cell in column_b:
        print(cell.value)
        list = f"{list + str(cell.value)} \n"
    label_b.config(text=list)


ba = Button(root, text="get column A", command=get_a)
ba.pack()

bb = Button(root, text="get column A", command=get_b)
bb.pack()

label_a = Label(root, text="")
label_a.pack()

label_b = Label(root, text="")
label_b.pack()

ws["A8"] = "Fred"
ws["B8"] = "Cheese"

wb.save("Pizza2.xlsx")

root.mainloop()
